
# coding UTF-8
# author:huangZengli
from osgeo import gdal
from pylab import *  # 支持中文

mpl.rcParams['font.sans-serif'] = ['SimHei']
from openpyxl import Workbook

# 创建一个Workbook对象
work = Workbook()

def out(data, name):
    ws = work.active
    ws['A1'] = '经度'
    ws['B1'] = '纬度'
    ws['C1'] = '高程'
    ws['D1'] = '所在栅格行'
    ws['E1'] = '所在栅格列'
    for i in range(len(data)):
        rows = []
        row_length = len(data[i])
        if row_length != 0:
            for j in range(row_length):
                rows.append(data[i][j])
                ws.append(rows[j])
        print(rows)
    work.save(name)

if __name__ == "__main__":
    filePath = 'test_jingweid11111111111111u.tif'  # tif文件路径
    dataset = gdal.Open(filePath)  # 打开tif

    # 获取行数列数和地理信息
    # geo_information(0):左上像素左上角的x坐标。
    # geo_information(1):w - e像素分辨率 / 像素宽度。
    # geo_information(2):行旋转（通常为零）。
    # geo_information(3):左上像素左上角的y坐标。
    # geo_information(4):列旋转（通常为零）。
    # geo_information(5):n - s像素分辨率 / 像素高度（北半球上图像为负值）
    geo_information = dataset.GetGeoTransform()
    col = dataset.RasterXSize  # 438
    row = dataset.RasterYSize  # 671
    band = dataset.RasterCount
    dem = dataset.GetRasterBand(1).ReadAsArray()
    # 获取行列数，对应其经纬度,j对于x坐标
    cols = []
    for y in range(row):  # 行
        rows = []
        for x in range(col):  # 列
            # 有效高程
            if dem[y][x] > 0:
                # 输出经纬度
                lon = geo_information[0] + x * geo_information[1] + y * geo_information[2]
                lat = geo_information[3] + x * geo_information[4] + y * geo_information[5]
                child = [lon, lat, dem[y][x], y, x]
                rows.append(child)
        cols.append(rows)
    out(cols, 'guandezhongwei1111111.xlsx')
    print('表已经生成')